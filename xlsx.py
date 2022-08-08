# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
import sys

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def updateWb(Infile,Outfile):
    print("xlsx:",Infile)
    wb = xl.load_workbook(Infile)
    sheet = wb['Sheet1']

    for row in range(3, sheet.max_row+1):
        cell1 = sheet.cell(row,4)
        cell2 = cell1.value * 0.9
        sheet.cell(row,5).value = cell2

    vals = Reference(sheet,
              min_row=3,
              max_row=sheet.max_row+2,
              min_col=4,max_col=4)

    chart = BarChart()
    chart.add_data(vals)
    sheet.add_chart(chart,'f2')
    wb.save(Outfile)

